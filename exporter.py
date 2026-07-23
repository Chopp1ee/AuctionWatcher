# exporter.py
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
from api import format_date, translate_status

def export_history_to_excel(chat_id, auction_id, history_data, auction_details):
    """Експортує історію змін в Excel файл"""
    
    # Створюємо нову книгу
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Історія змін"
    
    # Стилі
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ============ ІНФОРМАЦІЯ ПРО АУКЦІОН ============
    row = 1
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = f"ІСТОРІЯ ЗМІН АУКЦІОНУ"
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1a3c5e", end_color="1a3c5e", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    row = 3
    # ID аукціону (повністю)
    ws[f'A{row}'] = "🆔 ID аукціону:"
    ws[f'B{row}'] = auction_details[0] if auction_details else "Невідомо"
    ws.merge_cells(f'B{row}:E{row}')
    ws[f'B{row}'].font = Font(bold=True, size=11)
    
    row = 4
    # Назва аукціону (повністю)
    ws[f'A{row}'] = "📄 Назва:"
    ws[f'B{row}'] = auction_details[1] if auction_details and len(auction_details) > 1 else "Невідомо"
    ws.merge_cells(f'B{row}:E{row}')
    ws[f'B{row}'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    row = 5
    # Поточний статус
    ws[f'A{row}'] = "📊 Поточний статус:"
    status_text = translate_status(auction_details[2]) if auction_details and len(auction_details) > 2 else "Невідомо"
    ws[f'B{row}'] = status_text
    ws[f'C{row}'] = "📅 Дата додавання:"
    ws[f'D{row}'] = format_date(auction_details[4]) if auction_details and len(auction_details) > 4 else "Невідомо"
    
    row = 6
    ws[f'A{row}'] = "🕒 Останнє оновлення:"
    ws[f'B{row}'] = format_date(auction_details[3]) if auction_details and len(auction_details) > 3 else "Невідомо"
    
    # Порожній рядок перед таблицею
    row = 8
    
    # ============ ЗАГОЛОВКИ ТАБЛИЦІ ============
    headers = ["№", "Стара дата", "Нова дата"]
    header_cols = ['A', 'B', 'C']
    
    # Налаштовуємо ширину колонок
    ws.column_dimensions['A'].width = 22   # № - ширше (200 пікселів приблизно)
    ws.column_dimensions['B'].width = 25   # Стара дата
    ws.column_dimensions['C'].width = 25   # Нова дата
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    
    for col, header in zip(header_cols, headers):
        cell = ws[f'{col}{row}']
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # ============ ДАНІ ІСТОРІЇ ============
    data_row = row + 1
    for idx, (old_status, new_status, old_date, new_date, changed_at) in enumerate(history_data, 1):
        # № - номер зміни
        ws[f'A{data_row}'] = idx
        ws[f'A{data_row}'].border = border
        ws[f'A{data_row}'].alignment = center_alignment
        
        # Стара дата
        ws[f'B{data_row}'] = format_date(old_date)
        ws[f'B{data_row}'].border = border
        ws[f'B{data_row}'].alignment = center_alignment
        
        # Нова дата
        ws[f'C{data_row}'] = format_date(new_date)
        ws[f'C{data_row}'].border = border
        ws[f'C{data_row}'].alignment = center_alignment
        
        data_row += 1
    
    # ============ ФІКСУЄМО ЗАГОЛОВКИ ============
    ws.freeze_panes = 'A9'
    
    # ============ ІНФОРМАЦІЯ ВНИЗУ ============
    last_row = data_row
    ws[f'A{last_row + 2}'] = f"📊 Всього змін: {len(history_data)}"
    ws[f'A{last_row + 2}'].font = Font(bold=True, size=11)
    ws.merge_cells(f'A{last_row + 2}:E{last_row + 2}')
    
    ws[f'A{last_row + 3}'] = f"📅 Експорт створено: {format_date(datetime.now().isoformat())}"
    ws.merge_cells(f'A{last_row + 3}:E{last_row + 3}')
    
    # ============ НАЛАШТУВАННЯ ВИСОТИ РЯДКІВ ============
    # Висота рядка для назви (щоб влізла повністю)
    if auction_details and len(auction_details) > 1 and auction_details[1]:
        title_length = len(auction_details[1])
        if title_length > 50:
            ws.row_dimensions[4].height = 40
        elif title_length > 30:
            ws.row_dimensions[4].height = 30
        else:
            ws.row_dimensions[4].height = 20
    
    # Автоматична висота для рядків з даними
    for r in range(9, data_row):
        ws.row_dimensions[r].height = 25
    
    # Зберігаємо файл
    filename = f"history_{auction_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join("exports", filename)
    
    # Створюємо папку для експортів
    if not os.path.exists("exports"):
        os.makedirs("exports")
    
    wb.save(filepath)
    
    return filepath, filename

def cleanup_old_exports(days=7):
    """Видаляє старі експорти (за замовчуванням 7 днів)"""
    export_dir = "exports"
    if not os.path.exists(export_dir):
        return
    
    now = datetime.now()
    for filename in os.listdir(export_dir):
        if filename.endswith('.xlsx'):
            filepath = os.path.join(export_dir, filename)
            file_time = datetime.fromtimestamp(os.path.getctime(filepath))
            if (now - file_time).days > days:
                os.remove(filepath)